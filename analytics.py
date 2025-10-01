import os
import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from config import DB_CONFIG

os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

conn = psycopg2.connect(**DB_CONFIG)

queries = {
    "line_chart": """
        SELECT DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
               COUNT(*) AS num_orders
        FROM olist_orders_dataset o
        JOIN olist_order_items_dataset oi ON o.order_id = oi.order_id
        JOIN olist_customers_dataset c ON o.customer_id = c.customer_id
        GROUP BY month
        ORDER BY month;
    """,

    "bar_chart": """
        SELECT p.product_category_name AS category,
               SUM(oi.price) AS total_revenue
        FROM olist_order_items_dataset oi
        JOIN olist_products_dataset p ON oi.product_id = p.product_id
        JOIN olist_orders_dataset o ON oi.order_id = o.order_id
        GROUP BY p.product_category_name
        ORDER BY total_revenue DESC
        LIMIT 10;
    """,

    "barh_chart": """
        SELECT s.seller_state,
               AVG(oi.freight_value) AS avg_freight
        FROM olist_order_items_dataset oi
        JOIN olist_sellers_dataset s ON oi.seller_id = s.seller_id
        JOIN olist_orders_dataset o ON oi.order_id = o.order_id
        GROUP BY s.seller_state
        ORDER BY avg_freight DESC
        LIMIT 10;
    """,

    "pie_chart": """
        SELECT payment_type,
               COUNT(*) AS count
        FROM olist_order_payments_dataset p
        JOIN olist_orders_dataset o ON p.order_id = o.order_id
        JOIN olist_order_items_dataset oi ON o.order_id = oi.order_id
        GROUP BY payment_type;
    """,

    "hist_chart": """
        SELECT review_score
        FROM olist_order_reviews_dataset r
        JOIN olist_orders_dataset o ON r.order_id = o.order_id
        JOIN olist_order_items_dataset oi ON o.order_id = oi.order_id;
    """,

    "scatter_chart": """
        SELECT oi.price, oi.freight_value
        FROM olist_order_items_dataset oi
        JOIN olist_orders_dataset o ON oi.order_id = o.order_id
        JOIN olist_sellers_dataset s ON oi.seller_id = s.seller_id;
    """,

    "slider_chart": """
        SELECT DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
               s.seller_state,
               COUNT(*) AS num_orders
        FROM olist_orders_dataset o
        JOIN olist_order_items_dataset oi ON o.order_id = oi.order_id
        JOIN olist_sellers_dataset s ON oi.seller_id = s.seller_id
        GROUP BY month, s.seller_state
        ORDER BY month, s.seller_state;
    """
}

dataframes = {}

for chart_type, query in queries.items():
    df = pd.read_sql(query, conn)
    dataframes[chart_type] = df
    print(f"[{chart_type}] → {len(df)} rows")

    if chart_type == "line_chart":
        df.set_index('month')['num_orders'].plot.line(title="Number of Orders Per Month")
        plt.xlabel("Month")
        plt.ylabel("Number of Orders")
        plt.tight_layout()
        plt.savefig("charts/line_chart.png")
        plt.clf()

    elif chart_type == "bar_chart":
        ax = df.plot.bar(x='category', y='total_revenue', title="Top 10 Product Categories by Revenue", legend=False)
        plt.xlabel("Category")
        plt.ylabel("Total Revenue (in millions)")
        ax.get_yaxis().set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x/1e6:.1f}M'))
        for i, value in enumerate(df['total_revenue']):
            ax.text(i, value + 0.01 * value, f'{value/1e6:.1f}M', ha='center', va='bottom', fontsize=8)
        plt.tight_layout()
        plt.savefig("charts/bar_chart.png")
        plt.clf()

    elif chart_type == "barh_chart":
        df.set_index('seller_state')['avg_freight'].plot.barh(title="Average Freight Cost by Seller State")
        plt.xlabel("Average Freight Cost")
        plt.ylabel("Seller State")
        plt.tight_layout()
        plt.savefig("charts/barh_chart.png")
        plt.clf()

    elif chart_type == "pie_chart":
        df.set_index('payment_type')['count'].plot.pie(autopct='%1.1f%%', title="Payment Method Distribution")
        plt.ylabel("")
        plt.tight_layout()
        plt.savefig("charts/pie_chart.png")
        plt.clf()

    elif chart_type == "hist_chart":
        df['review_score'].plot.hist(bins=5, title="Distribution of Review Scores")
        plt.xlabel("Review Score")
        plt.ylabel("Count")
        plt.tight_layout()
        plt.savefig("charts/hist_chart.png")
        plt.clf()

    elif chart_type == "scatter_chart":
        df.plot.scatter(x='price', y='freight_value', title="Product Price vs Freight Cost")
        plt.xlabel("Product Price")
        plt.ylabel("Freight Cost")
        plt.tight_layout()
        plt.savefig("charts/scatter_chart.png")
        plt.clf()

slider_df = dataframes["slider_chart"]
slider_df['month'] = slider_df['month'].dt.strftime('%Y-%m')

fig = px.bar(
    slider_df,
    x="seller_state",
    y="num_orders",
    color="seller_state",
    animation_frame="month",
    title="Orders by Seller State Over Time"
)
fig.update_layout(xaxis_title="Seller State", yaxis_title="Number of Orders")
fig.show()

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def export_to_excel(dataframes_dict, filename):
    filepath = os.path.join("exports", filename)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            if sheet_name == "slider_chart":
                continue
            df_copy = df.copy()
            if 'month' in df_copy.columns:
                df_copy['month'] = pd.to_datetime(df_copy['month']).dt.strftime('%Y-%m')

            df_copy.to_excel(writer, index=False, sheet_name=sheet_name)


        workbook = writer.book

        for sheet_name in dataframes_dict:
            if sheet_name == "slider_chart":
                continue
            ws = workbook[sheet_name]
            ws.freeze_panes = "B2"
            ws.auto_filter.ref = ws.dimensions

            df = dataframes_dict[sheet_name]
            for col_idx, col in enumerate(df.columns):
                if pd.api.types.is_numeric_dtype(df[col]):
                    col_letter = get_column_letter(col_idx + 1)
                    max_row = len(df) + 1
                    rule = ColorScaleRule(
                        start_type="min", start_color="FFAA0000",
                        mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                        end_type="max", end_color="FF00AA00"
                    )
                    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{max_row}", rule)

    total_rows = sum(len(df) for name, df in dataframes_dict.items() if name != "slider_chart")
    print(f"Created file {filename}, {len(dataframes_dict)-1} sheets, {total_rows} rows.")

export_to_excel(dataframes, "export.xlsx")

conn.close()
